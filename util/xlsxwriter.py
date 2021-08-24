# coding:utf-8
import openpyxl


class XlsxWriter:
    def __init__(self, listHeaders):
        # 构造失败抛出异常
        self.wb_ = openpyxl.workbook.Workbook()
        self.listHeaders_ = listHeaders             # 记录header
        self.writeRow_ = 1                          # 当前写入的行位置
        self.ws_ = self.wb_.active

    def __del__(self):
        self.wb_.close()

    def save(self, path):
        self.wb_.save(path)

    def writeHeader(self):
        for col in range(len(self.listHeaders_)):
            c = col + 1
            self.ws_.cell(row=self.writeRow_, column=c).value = self.listHeaders_[col]
        self.writeRow_ += 1

    def writeData(self, dictData):
        col = 1
        for header in self.listHeaders_:
            if header in dictData.keys():
                self.ws_.cell(row=self.writeRow_, column=col).value = dictData[header]
            else:
                self.ws_.cell(row=self.writeRow_, column=col).value = "None"
            col += 1
        self.writeRow_ += 1

