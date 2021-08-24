# coding:utf-8
import openpyxl


class XlsxReader:
    def __init__(self, path):
        # 构造失败抛出异常
        self.path_ = path
        self.dictStatistical = dict
        self.wb_ = openpyxl.load_workbook(self.path_)
        self.sheetNames_ = self.wb_.get_sheet_names()
        self.ws_ = None  # 记录当前打开的sheet
        self.sheetName_ = ""  # 记录当前打开的sheet名称
        self.sheetCurRow_ = 2  # 记录当前读取的sheet行数位置
        self.sheetMaxCols_ = 0  # 当前sheet的最大列数
        self.sheetMaxRows_ = 0  # 当前sheet的最大行数

    def __del__(self):
        self.wb_.close()

    def getSheetNames(self):
        return self.sheetNames_

    def openSheet(self, sheetName):
        self.closeSheet()
        self.ws_ = self.wb_.get_sheet_by_name(sheetName)
        self.sheetName_ = sheetName
        self.sheetMaxRows_ = self.ws_.max_row
        self.sheetMaxCols_ = self.ws_.max_column
        self.sheetCurRow_ = 2

    def closeSheet(self):
        self.ws_ = None
        self.sheetName_ = ""
        self.sheetMaxRows_ = 0
        self.sheetMaxCols_ = 0
        self.sheetCurRow_ = 2

    def getSheetHeader(self, sheetName):
        if self.ws_ is None:
            return []
        cols = self.sheetMaxCols_
        listHeaders = []
        for col in range(1, cols + 1):
            value = str(self.ws_.cell(row=1, column=col).value).strip()
            listHeaders.append(value)
        return listHeaders

    # 读取一条记录, 默认从第二行开始, 第一行酸列头
    def getOneRecord(self, listHeaders):
        if self.ws_ is None:
            return {}

        if self.sheetCurRow_ > self.sheetMaxRows_:
            return {}

        dictRecord = {}
        for col in range(1, self.sheetMaxCols_ + 1):
            value = str(self.ws_.cell(row=self.sheetCurRow_, column=col).value).strip()
            dictRecord[listHeaders[col - 1]] = value
        dictRecord["数据来源"] = "文件:{}, sheet:{}, 行数:{}".format(self.path_, self.sheetName_, self.sheetCurRow_)
        self.sheetCurRow_ += 1
        return dictRecord
